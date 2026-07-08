import csv
import io
import os
import re
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field

from django.contrib.auth.hashers import make_password
from django.db import transaction
from openpyxl import load_workbook

from accounts.models import User, UserRole

ALLOWED_EXTENSIONS = {".csv", ".xlsx"}
MAX_FILE_SIZE_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 10_000
CPM_MAX_LENGTH = 50
MC_MAX_LENGTH = 100
EXISTING_LOOKUP_CHUNK = 2_000
BULK_CREATE_BATCH = 1_000
PASSWORD_HASH_WORKERS = min(16, max(4, (os.cpu_count() or 4) * 2))

CPM_COLUMN_ALIASES = {
    "cpm number",
    "cpm_number",
    "cpm",
    "cpm no",
    "cpm no.",
    "cpmnumber",
}
MC_COLUMN_ALIASES = {
    "mc number",
    "mc_number",
    "mc",
    "mc no",
    "mc no.",
    "mcnumber",
    "password",
}

DUPLICATE_REASONS = {
    "duplicate_in_file": "Duplicate CPM Number in this file.",
    "already_exists": "CPM Number already exists in the database.",
}

CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")


@dataclass
class RowError:
    row: int
    cpm_number: str | None
    reason: str


@dataclass
class DuplicateEntry:
    row: int
    cpm_number: str
    reason: str


@dataclass
class ImportResult:
    total_rows: int = 0
    successful: int = 0
    failed_rows: list[RowError] = field(default_factory=list)
    duplicates: list[DuplicateEntry] = field(default_factory=list)


def _normalize_header(value) -> str:
    if value is None:
        return ""
    normalized = str(value).strip().lower()
    normalized = re.sub(r"[\s_-]+", " ", normalized)
    return normalized.strip()


def _normalize_cpm(value) -> str:
    return str(value).strip().upper()


def _cell_value(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"none", "null", "n/a", "-"}:
        return ""
    return text


def _decode_csv_bytes(raw: bytes) -> str:
    last_error: UnicodeDecodeError | None = None
    for encoding in CSV_ENCODINGS:
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error is not None:
        raise ValueError(
            "Could not read CSV file. Save it as UTF-8 or try exporting again from Excel."
        ) from last_error
    raise ValueError("Could not read CSV file.")


def _detect_csv_dialect(sample: str) -> csv.Dialect:
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        return csv.excel


# Parsed data row: (file row number, raw CPM value, raw MC value).
ParsedRow = tuple[int, str, str]


def _find_column_index(headers: list[str], aliases: set[str]) -> int | None:
    for index, header in enumerate(headers):
        if header in aliases:
            return index
    return None


def _cell_at(values, index: int | None) -> str:
    if index is None or index >= len(values):
        return ""
    return _cell_value(values[index])


def _parse_csv(file_obj) -> tuple[list[str], list[ParsedRow]]:
    raw = file_obj.read()
    if isinstance(raw, str):
        content = raw
    else:
        content = _decode_csv_bytes(raw)

    if not content.strip():
        raise ValueError("File is empty or missing a header row.")

    sample = content[:4096]
    dialect = _detect_csv_dialect(sample)
    reader = csv.reader(io.StringIO(content), dialect=dialect)

    try:
        header_cells = next(reader)
    except StopIteration as exc:
        raise ValueError("File is empty or missing a header row.") from exc

    headers = [_normalize_header(name) for name in header_cells]
    if not any(headers):
        raise ValueError("File is empty or missing a header row.")

    cpm_index = _find_column_index(headers, CPM_COLUMN_ALIASES)
    mc_index = _find_column_index(headers, MC_COLUMN_ALIASES)

    rows: list[ParsedRow] = []
    for row_number, cells in enumerate(reader, start=2):
        cpm_raw = _cell_at(cells, cpm_index)
        mc_raw = _cell_at(cells, mc_index)
        if not cpm_raw and not mc_raw:
            continue
        rows.append((row_number, cpm_raw, mc_raw))

    return headers, rows


def _parse_xlsx(file_obj) -> tuple[list[str], list[ParsedRow]]:
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=True)

        try:
            header_row = next(row_iter)
        except StopIteration as exc:
            raise ValueError("File is empty or missing a header row.") from exc

        headers = [_normalize_header(cell) for cell in header_row]
        cpm_index = _find_column_index(headers, CPM_COLUMN_ALIASES)
        mc_index = _find_column_index(headers, MC_COLUMN_ALIASES)

        rows: list[ParsedRow] = []
        for row_number, values in enumerate(row_iter, start=2):
            cpm_raw = _cell_at(values, cpm_index)
            mc_raw = _cell_at(values, mc_index)
            if not cpm_raw and not mc_raw:
                continue
            rows.append((row_number, cpm_raw, mc_raw))
    finally:
        workbook.close()

    return headers, rows


def _validate_columns(headers: list[str]) -> None:
    header_set = set(headers)
    missing = []
    if not header_set.intersection(CPM_COLUMN_ALIASES):
        missing.append("CPM Number")
    if not header_set.intersection(MC_COLUMN_ALIASES):
        missing.append("MC Number")
    if missing:
        labels = ", ".join(missing)
        raise ValueError(
            f"Missing required columns: {labels}. "
            "Your file must include CPM Number and MC Number columns."
        )


def _validate_cpm(cpm_number: str) -> str | None:
    if len(cpm_number) > CPM_MAX_LENGTH:
        return f"CPM Number must be {CPM_MAX_LENGTH} characters or fewer."
    if not re.fullmatch(r"[A-Z0-9][A-Z0-9 ./-]*", cpm_number):
        return "CPM Number contains invalid characters."
    return None


def _validate_mc(mc_number: str) -> str | None:
    if len(mc_number) > MC_MAX_LENGTH:
        return f"MC Number must be {MC_MAX_LENGTH} characters or fewer."
    if not mc_number:
        return "Missing MC Number."
    return None


def validate_import_file(uploaded_file) -> None:
    if uploaded_file.size > MAX_FILE_SIZE_BYTES:
        raise ValueError("File exceeds the 5 MB size limit.")

    name = uploaded_file.name.lower()
    if not any(name.endswith(ext) for ext in ALLOWED_EXTENSIONS):
        raise ValueError("Invalid file type. Only CSV and XLSX files are allowed.")


def parse_member_file(uploaded_file) -> tuple[list[str], list[ParsedRow]]:
    validate_import_file(uploaded_file)
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return _parse_csv(uploaded_file)
    if name.endswith(".xlsx"):
        return _parse_xlsx(uploaded_file)
    raise ValueError("Invalid file type. Only CSV and XLSX files are allowed.")


def _fetch_existing_cpms(cpm_numbers: list[str]) -> set[str]:
    if not cpm_numbers:
        return set()

    existing: set[str] = set()
    for start in range(0, len(cpm_numbers), EXISTING_LOOKUP_CHUNK):
        chunk = cpm_numbers[start : start + EXISTING_LOOKUP_CHUNK]
        existing.update(
            User.objects.filter(cpm_number__in=chunk).values_list("cpm_number", flat=True)
        )
    return existing


def _hash_mc_numbers(mc_numbers: list[str]) -> list[str]:
    if not mc_numbers:
        return []

    with ThreadPoolExecutor(max_workers=PASSWORD_HASH_WORKERS) as executor:
        return list(executor.map(make_password, mc_numbers))


def _bulk_create_members(users: list[User]) -> int:
    if not users:
        return 0

    created = 0
    with transaction.atomic():
        for start in range(0, len(users), BULK_CREATE_BATCH):
            batch = users[start : start + BULK_CREATE_BATCH]
            User.objects.bulk_create(batch, batch_size=BULK_CREATE_BATCH)
            created += len(batch)
    return created


def import_members(uploaded_file, academic_year: str) -> ImportResult:
    headers, rows = parse_member_file(uploaded_file)
    _validate_columns(headers)

    if len(rows) > MAX_IMPORT_ROWS:
        raise ValueError(f"File exceeds the maximum of {MAX_IMPORT_ROWS:,} data rows.")

    result = ImportResult(total_rows=len(rows))
    seen_cpms: dict[str, int] = {}
    valid_rows: list[tuple[int, str, str]] = []

    for index, cpm_raw, mc_raw in rows:
        if not cpm_raw:
            result.failed_rows.append(
                RowError(row=index, cpm_number=None, reason="Missing CPM Number.")
            )
            continue

        if not mc_raw:
            result.failed_rows.append(
                RowError(row=index, cpm_number=_normalize_cpm(cpm_raw), reason="Missing MC Number.")
            )
            continue

        cpm_number = _normalize_cpm(cpm_raw)
        cpm_error = _validate_cpm(cpm_number)
        if cpm_error:
            result.failed_rows.append(
                RowError(row=index, cpm_number=cpm_number, reason=cpm_error)
            )
            continue

        mc_error = _validate_mc(mc_raw)
        if mc_error:
            result.failed_rows.append(
                RowError(row=index, cpm_number=cpm_number, reason=mc_error)
            )
            continue

        if cpm_number in seen_cpms:
            result.duplicates.append(
                DuplicateEntry(
                    row=index,
                    cpm_number=cpm_number,
                    reason=DUPLICATE_REASONS["duplicate_in_file"],
                )
            )
            continue

        seen_cpms[cpm_number] = index
        valid_rows.append((index, cpm_number, mc_raw))

    if not valid_rows:
        return result

    cpm_numbers = [cpm for _, cpm, _ in valid_rows]
    existing_cpms = _fetch_existing_cpms(cpm_numbers)

    pending_users: list[tuple[int, str, str]] = []
    for row_num, cpm_number, mc_number in valid_rows:
        if cpm_number in existing_cpms:
            result.duplicates.append(
                DuplicateEntry(
                    row=row_num,
                    cpm_number=cpm_number,
                    reason=DUPLICATE_REASONS["already_exists"],
                )
            )
            continue

        pending_users.append((row_num, cpm_number, mc_number))

    if pending_users:
        mc_numbers = [mc_number for _, _, mc_number in pending_users]
        hashed_passwords = _hash_mc_numbers(mc_numbers)
        users_to_create = [
            User(
                cpm_number=cpm_number,
                mc_number=mc_number,
                password=hashed_password,
                role=UserRole.MEMBER,
                academic_year=academic_year,
                is_active=True,
            )
            for (_, cpm_number, mc_number), hashed_password in zip(
                pending_users, hashed_passwords, strict=True
            )
        ]
        result.successful = _bulk_create_members(users_to_create)

    return result
